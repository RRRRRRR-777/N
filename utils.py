# moduleをインポート
import datetime
import glob
import math
import os
import re
import shutil
import sys
import time
from datetime import timedelta
from logging import getLogger, StreamHandler, Formatter
from statistics import mean
from dotenv import load_dotenv

import matplotlib.cm as cm
import numpy as np
import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from matplotlib.colors import TwoSlopeNorm
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side
from selenium import webdriver
from selenium.webdriver.chrome import service as fs
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


"""
初期実行
"""


class InitProcess:
    # ログの出力名を設定
    def set_log(self):
        self.logger = getLogger(__name__)
        load_dotenv()  # .envファイルから値を読み込む
        log_level = int(os.getenv('output_logger_level'))
        # ログレベルを設定
        self.logger.setLevel(log_level)
        # ログをコンソール出力するための設定
        sh = StreamHandler()
        fmt = Formatter(
            '%(asctime)s.%(msecs)-3d [%(levelname)-7s][%(name)s] %(message)s', datefmt='%H:%M:%S')
        sh.setFormatter(fmt)
        self.logger.addHandler(sh)
        return self.logger

    # Seleniumの設定
    def set_selenium(self):
        options = webdriver.ChromeOptions()
        # ブラウザのオプションを設定
        # ChromeDriver でよく使うオプションのチートシート https://qiita.com/kawagoe6884/items/cea239681bdcffe31828
        dt_now = datetime.datetime.now()
        date = dt_now.strftime('%y%m%d')
        downloadDir = os.getcwd() + f"/StockData" + date
        options.add_experimental_option(
            "prefs", {"download.default_directory": downloadDir})  # ファイルの保存先を指定
        options.add_argument('--headless')  # ヘッドレスモードでブラウザを起動
        options.add_argument('--ignore-ssl-errors')  # SSL証明書処理
        # SSL認証(この接続ではプライバシーが保護されません)を無効
        options.add_argument('--ignore-certificate-errors')
        options.add_argument(
            '--blink-settings=imagesEnabled=false')  # 画像を非表示にする
        # ディスクのメモリスペースを使う。DockerやGcloudのメモリ対策でよく使われる
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--mute-audio')  # すべてのオーディオをミュートする
        options.add_argument(
            '--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.3 Safari/605.1.15')  # ユーザーエージェントの指定
        # webdriverをFalseにする
        options.add_argument('--disable-blink-features=AutomationControlled')
        # ドライバの設定
        self.driver = webdriver.Chrome(options=options)
        self.driver.implicitly_wait(10)  # 暗黙的な待機
        self.driver.set_window_size(1200, 1000)
        # headlessモードでファイルをダウンロードする際の追加設定
        self.driver.command_executor._commands["send_command"] = (
            'POST',
            '/session/$sessionId/chromium/send_command'
        )
        self.driver.execute(
            'send_command',
            params={
                'cmd': 'Page.setDownloadBehavior',
                'params': {'behavior': 'allow', 'downloadPath': downloadDir}
            }
        )
        return self.driver

    def execute(self):
        # 保存ディレクトリの設定
        dt_now = datetime.datetime.now()
        date = dt_now.strftime('%y%m%d')
        # 個別銘柄のデータの保存先
        csv_downloadDir = os.getcwd() + f"/StockData" + date
        try:
            p = glob.glob(
                os.getcwd() + f"/StockData*", recursive=True)[0]
            shutil.rmtree(p)
            self.logger.info(f"remove {p}")
        except:
            pass
        os.mkdir(csv_downloadDir)
        # CSVデータの保存先
        stock_downloadDir = os.getcwd() + f"/CsvData" + date
        try:
            p = glob.glob(os.getcwd() + f"/CsvData*",
                          recursive=True)[0]
            shutil.rmtree(p)
            self.logger.info(f"remove {p}")
        except:
            pass
        os.mkdir(stock_downloadDir)


"""
finvizから銘柄の配列を取得
"""


class PickFinviz:
    def __init__(self, logger):
        self.logger = logger

    def execute(self):
        # URL
        load_dotenv()  # .envファイルから値を読み込む
        url = os.getenv('finviz_url')
        # ページ数(num*20銘柄)
        num = 100
        # ファイル名
        out_path = os.path.join(glob.glob(
            os.getcwd()+"/StockData*", recursive=True)[0], "input.txt")

        with open(out_path, "w", encoding="utf-8") as file:
            # 1ページごとでループする
            for i in range(num):
                page = str(i * 20 + 1)
                site = requests.get(
                    url + page, headers={'User-Agent': 'Custom'})
                data = BeautifulSoup(site.text, 'html.parser')

                tr_tag = data.find_all("tr", {
                                       "class": "styled-row is-hoverable is-bordered is-rounded is-striped has-color-text"})
                # 1銘柄ごとループする
                for j in range(0, len(tr_tag), 1):
                    a_tag = [a.text for a in tr_tag[j].find_all("a")]
                    # 行をテキストファイルに書き込む (波線で区切る)
                    file.write("~".join(a_tag[1:]) + "\n")
                print(f"Done {(i+1) * 20}", end="\r")

            self.logger.info(f"---Done Write input.txt (PickFinviz)---")


"""
ヒストリカルデータをダウンロードする
"""


class HistData:
    def __init__(self, logger, driver):
        self.logger = logger
        self.driver = driver

    def execute(self):
        # ヒストリカルデータの保存先
        dt_now = datetime.datetime.now()
        date = dt_now.strftime('%y%m%d')
        downloadDir = os.getcwd() + f"/StockData" + date
        # 銘柄データの保存先
        stocksDir = glob.glob(
            os.getcwd() + f"/StockData*/input.txt", recursive=True)[0]

        # 銘柄の箱
        symbol = np.full((5000), 0, dtype=object)

        # データ期間の指定（st:開始、ed:終了）
        st = datetime.date(1970, 1, 1)
        ed = datetime.date.today()
        dt = datetime.date(1970, 1, 1)
        st = st - dt
        ed = ed - dt
        st = (st.days) * 86400
        ed = (ed.days) * 86400

        # inputファイルから銘柄群のシンボルを取得
        i = 0
        if (os.path.exists(stocksDir)):
            with open(stocksDir, 'r', encoding='shift-jis') as f:
                for line in f.readlines():
                    i += 1
                    toks = line.split('~')
                    symbol[i] = toks[0]
        # 銘柄数を記録
        nsym = i

        # IXICのヒストリカルデータをダウンロード
        load_dotenv()  # .envファイルから値を読み込む
        base_url = os.getenv('yfinance_nasdaq_url')
        url = base_url + str(st)+'&period2='+str(ed) + \
            '&interval=1d&events=history&includeAdjustedClose=true'

        # 日足データのダウンロード
        self.driver.get(url)
        # 2秒間の一時停止
        time.sleep(1)

        # 銘柄数の分だけループ
        load_dotenv()  # .envファイルから値を読み込む
        base_url = os.getenv('yfinance_stocks_url')
        for i in range(1, nsym+1):
            url = base_url+str(symbol[i])+'?period1=' + str(st)+'&period2='+str(ed) + \
                '&interval=1d&events=history&includeAdjustedClose=true'
            # 日足データのダウンロード
            try:
                self.driver.get(url)
            except Exception as e:
                self.logger.warning(f"{str(symbol[i])} {int(i)}\n{e}")
            print(f"{int(i)}/{int(nsym)} ", end="\r")

            # 1秒間の一時停止
            time.sleep(1.0)

        self.logger.info(f"---Done Download HistData (HistData)---")


"""
IXICのヒストリカルデータの列を増やす
"""


class ProcessNASDAQ:
    def __init__(self, logger):
        self.logger = logger

    def execute(self):
        # IXICのCSVを読み込む
        IXICdir = glob.glob(
            os.getcwd() + f"/StockData*/^IXIC.csv", recursive=True)[0]
        df = pd.read_csv(IXICdir)

        # 追加する列
        # 50日移動平均線
        df['SMA50'] = df['Adj Close'].rolling(50).mean()
        # 150日移動平均線
        df['SMA150'] = df['Adj Close'].rolling(150).mean()
        # 200日移動平均線
        df['SMA200'] = df['Adj Close'].rolling(200).mean()
        # 200日移動平均線の20日平均値
        df['SMA200 mean 20days'] = df['SMA200'].rolling(20).mean()
        # 200日移動平均線の20日前の値
        df['SMA200 befor 20days'] = df['SMA200'].shift(20)
        # 200日移動平均線と現在の株価のギャップ
        df['SMA200 Gap'] = round(
            ((df['Adj Close'] - df['SMA200']) / abs(df['SMA200'])) * 100, 2)
        # 52週最高値
        # min_periodsを使用して1つ以上のデータがあった場合の最大値を求める
        df['52W High'] = df['Adj Close'].rolling(260, min_periods=1).max()
        # 52週最高値の25%以内
        df['52W High*0.75'] = df['52W High']*0.75
        # 52週最安値
        # min_periodsを使用して1つ以上のデータがあった場合の最小値を求める
        df['52W Low'] = df['Adj Close'].rolling(260, min_periods=1).min()
        # 52週最安値の30%以上
        df['52W Low*1.3'] = df['52W Low']*1.3
        # UpDownVolumeRatio
        # 前日と比較し株価が上昇していた日の出来高を'Up'に、下落していた日の出来高を'Down'に格納する
        df['Up'] = df.loc[df['Adj Close'].diff() > 0, 'Volume']
        df['Down'] = df.loc[df['Adj Close'].diff() <= 0, 'Volume']
        # 欠損値を0で埋める
        df = df.fillna(0)
        # 過去50営業日のうち株価が上昇した日の出来高を下落した日の出来高で割った数値
        df['U/D'] = df['Up'].rolling(50).sum() / df['Down'].rolling(50).sum()

        # ミネルビィ二のトレンドテンプレートのNo1〜No7までの列を作成
        df[['No1', 'No2', 'No3', 'No4', 'No5', 'No6', 'No7']] = 0

        # No1 現在の株価が150日と200日の移動平均線を上回っている。
        df.loc[(df['Adj Close'] > df['SMA150']) & (
            df['Adj Close'] > df['SMA200']), 'No1'] = int(1)
        # No2 150日移動平均線は200日移動平均線を上回っている。
        df.loc[df['SMA150'] > df['SMA200'], 'No2'] = int(1)
        # No3 200日移動平均線は少なくとも1か月、上昇トレンドにある。
        df.loc[df['SMA200 mean 20days'] >
               df['SMA200 befor 20days'], 'No3'] = int(1)
        # No4 50日移動平均線は150日移動平均線と200日移動平均線を上回っている。
        df.loc[(df['SMA50'] > df['SMA150']) & (
            df['SMA50'] > df['SMA200']), 'No4'] = int(1)
        # No5 現在の株価は50日移動平均線を上回っている。
        df.loc[df['Adj Close'] > df['SMA50'], 'No5'] = int(1)
        # No6 現在の株価は52週安値よりも、少なくとも30％高い。
        df.loc[df['Adj Close'] > df['52W Low*1.3'], 'No6'] = int(1)
        # No7 現在の株価は52週高値から少なくとも25％以内にある。
        df.loc[df['Adj Close'] > df['52W High*0.75'], 'No7'] = int(1)
        # No1~No7の合計値
        df['Total'] = df['No1'] + df['No2'] + df['No3'] + \
            df['No4'] + df['No5'] + df['No6'] + df['No7']

        # 買い条件1
        df['BuyFlg1'] = 0
        df.loc[df['Total'] >= 5, 'BuyFlg1'] = int(1)
        # df.loc[(df['Total'] == 5) & (df['Total'].shift(1) == 4), 'BuyFlg1'] = 1
        # 買い条件2
        df['BuyFlg2'] = 0
        df.loc[df['U/D'] >= 1, 'BuyFlg2'] = int(1)
        # 買い条件3
        df['BuyFlg3'] = 0
        df.loc[(df['Total'] >= 5) & (df['No1'] == 1) & (df['No4'] == 1)
               & (df['No5'] == 1) & (df['No6'] == 1), 'BuyFlg3'] = int(1)
        # 売り条件1
        df['SellFlg1'] = 0
        df.loc[df['Total'] <= 4, 'SellFlg1'] = int(1)
        # 売り条件2
        df['SellFlg2'] = 0
        df.loc[df['Adj Close'] < df['SMA200'], 'SellFlg2'] = int(1)

        # Totalの列を^IXIC Totalに変更する
        df.rename(columns={'Total': '^IXIC Total'}, inplace=True)
        # Csvの書き出し
        df.to_csv(IXICdir, index=False)

        self.logger.info(f"---Done Process ProcessNASDAQ (ProcessNASDAQ)---")


"""
個別株のヒストリカルデータの列を増やす
"""


class ProcessHistData:
    def __init__(self, logger):
        self.logger = logger

    def execute(self):
        # IXICのCSVを読み込む
        IXICdir = glob.glob(
            os.getcwd() + f"/StockData*/^IXIC.csv", recursive=True)[0]
        dfIXIC = pd.read_csv(IXICdir)
        # 各企業のヒストリカルデータを読み込む
        cnt = 0
        files = glob.glob(
            os.getcwd() + "/StockData*/*.csv", recursive=True)
        # Comprehensiveのファイルを除く
        try:
            compdir = os.path.join(
                glob.glob(os.getcwd()+"/CsvData*")[0], "Comprehensive.csv")
            files.remove(compdir)
        except:
            pass
        mnum = len(files)
        for file in files:
            # 入力CSV
            df = pd.read_csv(file)
            s = re.sub(os.getcwd() +
                       r"/StockData[0-9]{6}/|.csv", "", file)

            # 'IXIC Total'列が複数結合されるのを防ぐため(本実装時には何度も当ファイルを実行されることが無いため当コードは不要と思う)
            if not '^IXIC Total' in df.columns:
                # 入力CSVと^IXICのCSVを結合する
                df = pd.merge(
                    df, dfIXIC[['Date', '^IXIC Total']], on='Date', how='inner')

            # 追加する列
            # 8日指数平滑移動平均線
            df['EMA8'] = df['Adj Close'].ewm(8).mean()
            # 10日移動平均線
            df['SMA10'] = df['Adj Close'].rolling(10).mean()
            # 21日指数平滑移動平均線
            df['EMA21'] = df['Adj Close'].ewm(21).mean()
            # 50日移動平均線
            df['SMA50'] = df['Adj Close'].rolling(50).mean()
            # 150日移動平均線
            df['SMA150'] = df['Adj Close'].rolling(150).mean()
            # 200日移動平均線
            df['SMA200'] = df['Adj Close'].rolling(200).mean()
            # 200日移動平均線の20日平均値
            df['SMA200 mean 20days'] = df['SMA200'].rolling(20).mean()
            # 200日移動平均線の20日前の値
            df['SMA200 befor 20days'] = df['SMA200'].shift(20)
            # 8日指数平滑移動平均線と現在の株価のギャップ
            df['EMA8 Gap'] = round(
                ((df['Adj Close'] - df['EMA8']) / abs(df['EMA8'])) * 100, 2)
            # 200日移動平均線と現在の株価のギャップ
            # df['SMA200 Gap'] = df['Adj Close'] / df['SMA200']
            df['SMA200 Gap'] = round(
                ((df['Adj Close'] - df['SMA200']) / abs(df['SMA200'])) * 100, 2)
            # 出来高の50日移動平均線
            df['Volume SMA50'] = df['Volume'].rolling(50).mean()
            # 52週最高値
            # min_periodsを使用して1つ以上のデータがあった場合の最大値を求める
            df['52W High'] = round(
                df['High'].rolling(260, min_periods=1).max(), 2)
            # 52週最高値の25%以内
            df['52W High*0.75'] = df['52W High']*0.75
            # 52週最安値
            # min_periodsを使用して1つ以上のデータがあった場合の最小値を求める
            df['52W Low'] = df['Low'].rolling(260, min_periods=1).min()
            # 52週最安値の30%以上
            df['52W Low*1.3'] = df['52W Low']*1.3
            # UpDownVolumeRatio 過去50営業日のうち株価が上昇した日の出来高を下落した日の出来高で割った数値
            # 前日と比較し株価が上昇していた日の出来高を'Up'
            df['Up'] = df.loc[df['Adj Close'].diff() > 0, 'Volume']
            # 前日と比較し株価が下落していた日の出来高を'Down'に格納する
            df['Down'] = df.loc[df['Adj Close'].diff() <= 0, 'Volume']
            df = df.fillna(0)  # 欠損値を0で埋める
            df['U/D'] = round(df['Up'].rolling(50).sum() /
                              df['Down'].rolling(50).sum(), 3)
            # ATH
            df['ATH'] = round(df['High'].max(), 2)

            # ミネルビィ二のトレンドテンプレートのNo1〜No7までの列を作成
            df[['No1', 'No2', 'No3', 'No4', 'No5', 'No6', 'No7']] = 0

            # No1 現在の株価が150日と200日の移動平均線を上回っている。
            df.loc[(df['Adj Close'] > df['SMA150']) & (
                df['Adj Close'] > df['SMA200']), 'No1'] = int(1)
            # No2 150日移動平均線は200日移動平均線を上回っている。
            df.loc[df['SMA150'] > df['SMA200'], 'No2'] = int(1)
            # No3 200日移動平均線は少なくとも1か月、上昇トレンドにある。
            df.loc[df['SMA200 mean 20days'] >
                   df['SMA200 befor 20days'], 'No3'] = int(1)
            # No4 50日移動平均線は150日移動平均線と200日移動平均線を上回っている。
            df.loc[(df['SMA50'] > df['SMA150']) & (
                df['SMA50'] > df['SMA200']), 'No4'] = int(1)
            # No5 現在の株価は50日移動平均線を上回っている。
            df.loc[df['Adj Close'] > df['SMA50'], 'No5'] = int(1)
            # No6 現在の株価は52週安値よりも、少なくとも30％高い。
            df.loc[df['Adj Close'] > df['52W Low*1.3'], 'No6'] = int(1)
            # No7 現在の株価は52週高値から少なくとも25％以内にある。
            df.loc[df['Adj Close'] > df['52W High*0.75'], 'No7'] = int(1)
            # No1~No7の合計値
            df['Total'] = df['No1'] + df['No2'] + df['No3'] + \
                df['No4'] + df['No5'] + df['No6'] + df['No7']
            # 昨日のTotal
            df['Prev Total'] = df['Total'].shift(1)
            df['Prev Total'] = df['Total'].shift(1)
            df['Prev Total'] = df['Total'].shift(1)

            # 買い条件1
            df['BuyFlg1'] = 0
            df.loc[dfIXIC['^IXIC Total'] >= 2, 'BuyFlg1'] = int(1)
            # 買い条件2
            df['BuyFlg2'] = 0
            df.loc[df['U/D'] >= 1, 'BuyFlg2'] = int(1)
            # # 買い条件3
            # df['BuyFlg3'] = 0
            # df.loc[(df['Total'] == 5) & (df['Total'].shift(1) == 4), 'BuyFlg3'] = 1
            # 買い条件4
            df['BuyFlg4'] = 0
            df.loc[(df['Total'] >= 5) & (df['No1'] == 1) & (df['No4'] == 1) & (
                df['No5'] == 1) & (df['No6'] == 1), 'BuyFlg4'] = int(1)
            # 売り条件1
            df['SellFlg1'] = 0
            df.loc[df['Total'] <= 4, 'SellFlg1'] = int(1)
            # 売り条件2
            df['SellFlg2'] = 0
            df.loc[df['Adj Close'] < df['SMA200'], 'SellFlg2'] = int(1)
            # # 買い控え条件1
            # df['NotBuyFlg1'] = 0
            # df.loc[(df['Total'] == 5) & (df['No7'] == 1), 'NotBuyFlg1'] = int(1)

            # 8%で売る
            # df.loc[df['BuyPrice']/ df['Adj Close'] >= 1.08, 'SellFlg3'] = int(1)

            try:
                # 買っているフラッグ
                df['BuyingFlg'] = 0
                # df.loc[(df['BuyFlg1'] == 1) & (df['BuyFlg2'] == 1) & (df['BuyFlg3'] == 1) & (df['BuyFlg4'] == 1) , 'BuyingFlg'] = int(1)
                df.loc[(df['BuyFlg1'] == 1) & (df['BuyFlg2'] == 1) &
                       (df['BuyFlg4'] == 1), 'BuyingFlg'] = int(1)
                # # 本日買い条件にあったもののみフラグを立てる
                # df.loc[df['BuyingFlg'].shift(1) == 1, 'BuyingFlg'] = int(0)
                # 売ったフラッグ
                df['SelledFlg'] = 0
                df.loc[(df['BuyingFlg'].cumsum() >= 1) & (df['SellFlg1'] == 1) | (
                    df['SellFlg2'] == 1), 'SelledFlg'] = int(1)  # 該当行より上の行でBuyingFlgが立っていれば立てる
                df.loc[df['SelledFlg'].shift(1) == 1, 'SelledFlg'] = int(0)

                # 'BuyingFlg'または'SelledFlg'が1のものを選出する
                df_trade = df[(df['BuyingFlg'] == 1) | (df['SelledFlg'] == 1)]
                # 'BuyingFlg'と'SelledFlg'各々の1が続く行をすべて0に変換する
                df_trade.loc[df_trade['BuyingFlg'].shift(
                    1) == 1, 'BuyingFlg'] = int(0)
                df_trade.loc[df_trade['SelledFlg'].shift(
                    1) == 1, 'SelledFlg'] = int(0)
                # 'BuyingFlg'または'SelledFlg'の値が1であるもののみを選出する
                df_trade = df_trade[(df_trade['BuyingFlg'] == 1) | (
                    df_trade['SelledFlg'] == 1)]

                # # 買い値
                # df_trade['BuyPrice'] = 0
                # df_trade.loc[df_trade['BuyingFlg'] == 1, 'BuyPrice'] = df_trade['Adj Close']

                # dfの一番初めがSelledFlgで始まった場合はその行を消す
                if df_trade['SelledFlg'].iloc[0] == 1:
                    df_trade = df_trade.iloc[1:]

                # 利益率
                df_trade['Earn'] = 0
                # df_trade.loc[df_trade['SelledFlg'] == 1, 'Earn'] = (df_trade['Adj Close'] / df_trade['Adj Close'].shift(1) - 1) * 100
                df_trade['SelledFlg'] = df_trade['SelledFlg'].astype(
                    float).ffill()
                df_trade['Adj Close'] = df_trade['Adj Close'].astype(
                    float).ffill()
                df_trade.loc[df_trade['SelledFlg'] == 1,
                             'Earn'] = df_trade['Adj Close'] / df_trade['Adj Close'].shift(1)

                # 元のBuyingFlgとSelledFlgをすべて0にする
                df[['BuyingFlg', 'SelledFlg']] = 0

                # 利益率をdfに追加するために結合
                # 'Earn'列が複数結合されるのを防ぐため(本実装時には何度も当ファイルを実行されることが無いため当コードは不要と思う)
                if not 'Earn' in df.columns:
                    # 入力CSVとdf_tradeのCSVを結合する
                    df = pd.merge(
                        df, df_trade[['Date', 'Earn']], on='Date', how='outer').fillna(0)

                # dfにBuyingFlgとSelledFlgの値を結合する
                df = pd.merge(
                    df, df_trade[['Date', 'BuyingFlg', 'SelledFlg']], on='Date', how='outer')
                # 重複した列を削除し、列名を変更する
                df = df.drop(['BuyingFlg_x', 'SelledFlg_x'], axis=1).rename(
                    columns={'BuyingFlg_y': 'BuyingFlg', 'SelledFlg_y': 'SelledFlg'})

                # 総利益率
                # df['TotalEarn'] = df[df['Earn'] != 0]['Earn'].cumsum()
                df['TotalEarn'] = np.cumprod(df[df['Earn'] != 0]['Earn'])
                # 0の箇所を前の値で埋める
                df['TotalEarn'] = df['TotalEarn'].ffill()
                df = df.fillna(0)

                # 買い値
                df.loc[df['BuyingFlg'] == 1, 'BuyPrice'] = df['Adj Close']
                # 0の箇所を前の値で埋める
                df = df.ffill()

                # 空白を0で埋める
                df = df.fillna(0)

            # 取引履歴がない場合Empty DataFrameエラーが発生するのでその場合は2つの列を追加する
            except:
                df[['Earn', 'TotalEarn']] = float(0)

            df.to_csv(file, index=False)
            cnt += 1
            print(f"{cnt}/{mnum} ", end="\r")

        self.logger.info(
            f"---Done Process ProcessHistData (ProcessHistData)---")


"""
RSを計算する
"""


class CalculateRS:
    # 関数の定義
    def __init__(self, logger):
        self.logger = logger
    # RSの素点を計算する関数

    def period_perf(self, data, n):
        i = n
        # 現在の四半期の取引日数
        period = i*self.oneYear//4
        # 有効な価格データが見つからない場合は最も古いデータを使用する
        try:
            data.iloc[-period]
        except:
            period = 0

        period_price = data.iloc[-period]  # 指定された期間の終値
        latest_price = data.iloc[-1]  # 最新の終値
        # RSの素点の計算
        calc = ((latest_price - period_price) / period_price) * 100

        return calc

    # RSのランキングを作成
    def calculate_percentile_ranking(self, data):
        # 辞書の値をソートしてキーと共に取得
        sorted_data = sorted(data.items(), key=lambda x: x[1])
        n = len(sorted_data)
        rankings = {}
        rank = 1

        for i in range(n):
            key, value = sorted_data[i]
            if i > 0 and value != sorted_data[i - 1][1]:
                # 前のデータと値が異なる場合、新しいランクを設定
                rank = i + 1
            # ランキングを小数点として設定
            percentile_rank = (rank - 1) / (n - 1) * 99.99
            rankings[key] = round(percentile_rank, 2)  # ランクを小数点2桁まで丸める

        return rankings

    def execute(self):
        # プログラムの実行時間の計測
        start = time.time()
        # ディレクトリの定義
        # input.txtのディレクトリ
        inputDir = os.path.join(
            glob.glob(os.getcwd()+"/StockData*")[0], "input.txt")
        # 個別株の終値データ
        stock_dir = glob.glob(
            os.getcwd()+"/StockData*/*.csv", recursive=True)
        IXICdir = glob.glob(os.getcwd(
        ) + f"/StockData*/^IXIC.csv", recursive=True)[0]  # IXICのファイルを除く
        stock_dir.remove(IXICdir)
        try:
            compdir = os.path.join(glob.glob(os.getcwd(
            )+"/CsvData*")[0], "Comprehensive.csv")  # Comprehensiveのファイルを除く
            stock_dir.remove(compdir)
        except:
            pass
        # 出力先のディレクトリ
        outputDir = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "Comprehensive.csv")

        # テキストファイルをDataFrameに変換する
        data = []

        if os.path.exists(inputDir):
            with open(inputDir, 'r', encoding='shift-jis') as f:
                for line in f.readlines():
                    toks = line.strip().split('~')
                    data.append(toks)

        # df = pd.DataFrame(data, columns=["Ticker", "Company", "Sector", "Industry", "MarketCap", "P/E", "FwdP/E", "InsiderOwn", "EarningsDate", "Volume", "Price"])
        df = pd.DataFrame(data, columns=["Ticker", "Company", "Sector", "Industry",
                          "MarketCap", "ROE", "InsiderOwn", "EarningsDate", "Volume", "Price"])

        # RSランキングを作成する
        self.oneYear = 252  # 1年間の営業日
        rs = {}
        data = {}  # 各銘柄のティッカーコードとRSの素点を格納する辞書型の変数
        max_num = len(stock_dir)  # 銘柄数の最大値
        # col = ["Date", "ATH", "52W High", "U/D", "SMA10", "EMA21", "SMA50", "SMA200", "Volume SMA50", "BuyFlg1", "BuyFlg2", "BuyFlg3", "BuyFlg4", "BuyingFlg"] # 指定する列のリスト
        col = ["Date", "ATH", "52W High", "U/D", "^IXIC Total", "Total", "Prev Total", "No1", "No4", "No5", "No6", "SMA10", "EMA8",
               "EMA21", "SMA50", "SMA200", "EMA8 Gap", "SMA200 Gap", "Volume SMA50", "BuyFlg1", "BuyFlg2", "BuyFlg4", "BuyingFlg"]  # 指定する列のリスト

        for i, stock_path in enumerate(stock_dir):
            stock_data = pd.read_csv(stock_path)  # 銘柄を一つ選択
            ticker = re.sub(os.getcwd(
            ) + r"/StockData[0-9]{6}/|.csv", "", stock_path)  # ティッカーコードを抽出

            # RSの素点を作成
            stock = stock_data['Adj Close']
            rs[ticker] = 2 * self.period_perf(stock, 1) + self.period_perf(
                stock, 2) + self.period_perf(stock, 3) + self.period_perf(stock, 4)

            # ヒストリカルデータから指定した列の最新の値をdataに格納する
            last_row = stock_data.iloc[-1]
            data[ticker] = last_row[col]

        rankings = self.calculate_percentile_ranking(rs)
        df['RS'] = df['Ticker'].map(rankings)

        hist_df = pd.DataFrame(data).T
        hist_df.reset_index(inplace=True)
        hist_df.columns = ["Ticker"] + col

        df = pd.merge(df, hist_df, on="Ticker", how="left")
        df.to_csv(outputDir, index=False)

        # プログラムの実行時間の出力
        self.logger.info(time.time() - start)
        self.logger.info(f"---Done Process CalculateRS (CalculateRS)---")


"""
BuyingStock.csvを出力する
"""


class BuyingStock:
    def __init__(self, logger):
        self.logger = logger

    def execute(self):
        input_dir = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "Comprehensive.csv")
        input_df = pd.read_csv(input_dir)
        data = []
        for i in range(0, len(input_df)):
            stock = input_df.iloc[i]
            # if stock["BuyingFlg"] == 1:
            #     data.append(stock)
            if (stock['U/D'] >= 1) & (stock['Total'] >= 5) & (stock['No1'] == 1) & (stock['No4'] == 1) & (stock['No5'] == 1) & (stock['No6'] == 1) & (stock["RS"] >= 70) & (stock["Volume SMA50"] >= 500000) & (stock["SMA200 Gap"] >= 10):
                data.append(stock)

        df = pd.DataFrame(data)
        outputDir = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "BuyingStock.csv")
        df.to_csv(outputDir, index=False)


"""
CとAを取得する
"""


class CurrentAnnual:
    # 関数の定義
    def __init__(self, logger, driver):
        self.logger = logger
        self.driver = driver
    # XPATH要素の値を取得する

    def get_element(self, base_element, num):
        ''' args
        base_element: 欲しい要素のベースになるxpath
        num: 欲しい値の数
        '''
        list = []
        for i in range(num):
            part_element = base_element.replace('$', str(i+1))
            element = WebDriverWait(self.driver, 30).until(
                EC.presence_of_element_located((By.XPATH, part_element)))
            value = self.driver.execute_script(
                "return arguments[0].innerText;", element)
            list.append(value)  # 欲しい要素をリストに追加
        return list

    # アナリスト情報を取得
    def analysts_info(self, ticker_code):
        try:
            load_dotenv()  # .envファイルから値を読み込む
            base_url = os.getenv('analysts_url')
            url = base_url + f"{ticker_code}/analysis"
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            eps = soup.select('.tableContainer tbody tr')[1].select('td')
            revenue = soup.select('.tableContainer')[
                1].select('tbody tr')[1].select('td')
            self.analysts_data.append([ticker_code, eps[1].text, eps[2].text, eps[3].text,
                                      eps[4].text, revenue[1].text, revenue[2].text, revenue[3].text, revenue[4].text])
        except Exception as e:
            self.analysts_data.append(
                [ticker_code, None, None, None, None, None, None, None, None])
            self.logger.warning(
                f"Failed to fetch the page for {ticker_code} AnalystInfo. \n {e}")

    # 過去の情報を取得
    def previous_info(self, ticker_code):
        # ターゲットのURL
        load_dotenv()  # .envファイルから値を読み込む
        base_url = os.getenv('finance_url')
        url = base_url + f"{ticker_code}"
        try:
            # URLにアクセス
            self.driver.get(url)
            # Financialsタブを押下(デフォルトはAnnualタブ)
            financial_tab = WebDriverWait(self.driver, 30).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="result-tab-filings"][2]')))
            self.driver.execute_script("arguments[0].click();", financial_tab)
            # 年を取得
            year = self.get_element(
                '//*[@id="dashContentWrapper"]/div[3]/div[1]/div/div[2]/div/div[2]/div[2]/div[$]/div', 3)
            # AnnualEPS
            annual_eps = self.get_element(
                '//*[@id="dashContentWrapper"]/div[3]/div[1]/div/div[2]/div/div[2]/div[30]/div[$]/div', 3)
            for date, value in zip(year, annual_eps):
                self.previous_data.append(
                    [ticker_code, date, value, None, None, None, None])
            # AnnualRevenue
            annual_revenue = self.get_element(
                '//*[@id="dashContentWrapper"]/div[3]/div[1]/div/div[2]/div/div[2]/div[4]/div[$]/div', 3)
            for date, value in zip(year, annual_revenue):
                self.previous_data.append(
                    [ticker_code, date, None, value, None, None, None])

            # Quarterタブを押下
            quarter_tab = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="dashContentWrapper"]/div[2]/div[2]/div[2]/div')))
            self.driver.execute_script("arguments[0].click();", quarter_tab)
            # 四半期時期を取得
            quarter = self.get_element(
                '//*[@id="dashContentWrapper"]/div[3]/div[1]/div/div[2]/div/div[2]/div[2]/div[$]/div', 6)
            # QuarterEPS
            quarter_eps = self.get_element(
                '//*[@id="dashContentWrapper"]/div[3]/div[1]/div/div[2]/div/div[2]/div[30]/div[$]/div', 6)
            for date, value in zip(quarter, quarter_eps):
                self.previous_data.append(
                    [ticker_code, None, None, None, date, value, None])
            # QuarterRevenue
            quarter_revenue = self.get_element(
                '//*[@id="dashContentWrapper"]/div[3]/div[1]/div/div[2]/div/div[2]/div[4]/div[$]/div', 6)
            for date, value in zip(quarter, quarter_revenue):
                self.previous_data.append(
                    [ticker_code, None, None, None, date, None, value])

        except Exception as e:
            self.logger.warning(f"---Exception '{ticker_code}'---\n{e}")
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            self.logger.warning(f"{exc_type}, {fname}, {exc_tb.tb_lineno}row")

    # データが存在しないときの関数
    def try_exist_index(self, data, index):
        try:
            if 0 <= int(index) < len(data):
                return data.iloc[index]
            else:
                return np.nan
        except (ValueError, TypeError, IndexError):
            return np.nan

    # 浮動小数に変換できない場合の関数
    def try_float_data(self, data):
        try:
            return float(data)
        except:
            return np.nan

    # 対象の期間のパフォーマンスを計算
    def calculate(self, num1, num2, num1_index, num2_index):
        # 値の型を浮動小数点に変換
        num1 = self.try_float_data(self.try_exist_index(num1, num1_index))
        num2 = self.try_float_data(self.try_exist_index(num2, num2_index))
        if num2 == 0 or pd.isna(num2):
            return np.nan
        else:
            try:
                return round(((num1 - num2) / abs(num2)) * 100, 2)
            except:
                return np.nan

    # 桁数を変換
    def convert_value(self, value):
        if isinstance(value, float) or pd.isna(value):
            return value
        elif 'k' in value:
            return float(value.replace('k', '')) * 0.001
        elif 'M' in value:
            return float(value.replace('M', ''))
        elif 'B' in value:
            return float(value.replace('B', '')) * 1000
        else:
            return None

    def execute(self):
        # 入力するdfの設定
        stock_dir = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "BuyingStock.csv")
        stock_df = pd.read_csv(stock_dir)["Ticker"]
        self.analysts_data = []
        self.previous_data = []
        self.error_list = []
        pd.options.display.float_format = '{:.2f}'.format
        for i, ticker_code in enumerate(stock_df):
            self.analysts_info(ticker_code)
            time.sleep(1)  # 1秒間待機
            self.previous_info(ticker_code)
            time.sleep(1)  # 1秒間待機
            # # 5銘柄ごとに30秒間待機
            # if (i!=0)&(i%3==0):
            #     print(f"sleep ", end="\r")
            #     time.sleep(10)
            print(f"{i+1}/{len(stock_df)} ", end="\r")
        # Errorになった銘柄のデータを再度取得
        if len(self.error_list) != 0:
            self.logger.info("error_list Process")
            time.sleep(15)  # 待機時間
            self.logger.info(self.error_list)
            for i, ticker_code in enumerate(self.error_list):
                self.analysts_info(ticker_code)
                self.previous_info(ticker_code)
                # 1秒間待機
                time.sleep(1)
                print(f"{i+1}/{len(self.error_list)} ", end="\r")
                # 5銘柄ごとに30秒間待機
                if (i != 0) & (i % 5 == 0):
                    print(f"sleep ", end="\r")
                    time.sleep(30)
        # ブラウザを閉じる
        self.driver.quit()

        # 出力するdfの設定
        col = ["Ticker", "CQ_EPS", "NQ_EPS", "CY_EPS",
               "NY_EPS", "CQ_Rev", "NQ_Rev", "CY_Rev", "NY_Rev"]
        analysts_df = pd.DataFrame(self.analysts_data, columns=col)
        col = ["Ticker", "Year", "Annual EPS", "Annual Revenue",
               "Date", "Quarter EPS", "Quarter Revenue"]
        previous_df = pd.DataFrame(self.previous_data, columns=col)
        # dfの結合
        df = pd.concat([analysts_df, previous_df], axis=0, ignore_index=True)

        outputDir = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "TmpCurrentAnnual.csv")
        df.to_csv(outputDir, index=False)
        df = pd.read_csv(outputDir)

        self.logger.info("Calculating Data")

        # Revenue列の各セルに対して変換関数を適用
        df['CQ_Rev'] = df['CQ_Rev'].apply(self.convert_value)
        df['NQ_Rev'] = df['NQ_Rev'].apply(self.convert_value)
        df['CY_Rev'] = df['CY_Rev'].apply(self.convert_value)
        df['NY_Rev'] = df['NY_Rev'].apply(self.convert_value)

        # 日付の列のフォーマットを変更する
        df['Year'] = pd.to_datetime(df['Year'], format='%Y/%m/%d')
        df['Date'] = pd.to_datetime(df['Date'], format='%Y/%m/%d')

        # EPSの値を適応
        df['Annual EPS'] = df['Annual EPS'].replace(
            '[\$,]', '', regex=True).astype(float)
        df['Quarter EPS'] = df['Quarter EPS'].replace(
            '[\$,]', '', regex=True).astype(float)
        # Revenueの値を適応
        df['Annual Revenue'] = df['Annual Revenue'].replace(
            '[\$,]', '', regex=True).replace(',', '', regex=True)
        df['Quarter Revenue'] = df['Quarter Revenue'].replace(
            '[\$,]', '', regex=True).replace(',', '', regex=True)

        # 重複列を削除
        df = df[~df.duplicated(keep='first')]
        # 値が0の場合はNaNにする
        df.replace(0, np.nan, inplace=True)
        # 各値のYoYパフォーマンスの列を作成する
        data = []  # 作成データを格納するリスト

        for i, ticker_code in enumerate(stock_df):
            self.logger.debug(f"ticker: {ticker_code}")
            ticker_df = df[df["Ticker"] == ticker_code]
            # 四半期EPSを設定
            quarter_eps = ticker_df[pd.notna(ticker_df["Date"])].groupby(
                "Date").nth(0).sort_values("Date", ascending=False)["Quarter EPS"]
            # 各四半期EPSのパフォーマンスを計算
            perf_previous_quarter_eps_2 = self.calculate(
                num1=quarter_eps, num2=quarter_eps, num1_index=1, num2_index=5)
            perf_previous_quarter_eps = self.calculate(
                num1=quarter_eps, num2=quarter_eps, num1_index=0, num2_index=4)
            perf_courrent_quarter_eps = self.calculate(
                num1=ticker_df["CQ_EPS"].dropna(), num2=quarter_eps, num1_index=0, num2_index=3)
            perf_next_quarter_eps = self.calculate(
                num1=ticker_df["NQ_EPS"].dropna(), num2=quarter_eps, num1_index=0, num2_index=2)
            self.logger.debug(
                f"`EPS`\n2期前:{perf_previous_quarter_eps_2}, 1期前:{perf_previous_quarter_eps}, 今期:{perf_courrent_quarter_eps}, 来期:{perf_next_quarter_eps}\n")

            # 年度EPSを設定
            annual_eps = ticker_df[pd.notna(ticker_df["Year"])].groupby(
                "Year").nth(0).sort_values("Year", ascending=False)["Annual EPS"]
            # 年度EPSのパフォーマンスを計算
            perf_previous_annual_eps_2 = self.calculate(
                num1=annual_eps, num2=annual_eps, num1_index=1, num2_index=2)
            perf_previous_annual_eps = self.calculate(
                num1=annual_eps, num2=annual_eps, num1_index=0, num2_index=1)
            perf_courrent_annual_eps = self.calculate(
                num1=ticker_df["CY_EPS"].dropna(), num2=annual_eps, num1_index=0, num2_index=0)
            perf_next_annual_eps = self.calculate(
                num1=ticker_df["NY_EPS"].dropna(), num2=ticker_df["CY_EPS"].dropna(), num1_index=0, num2_index=0)
            self.logger.debug(
                f"`EPS`\n一昨年:{perf_previous_annual_eps_2}, 昨年:{perf_previous_annual_eps}, 今年:{perf_courrent_annual_eps}, 来年:{perf_next_annual_eps}\n")

            # 四半期Revenueを設定
            quarter_revenue = ticker_df[pd.notna(ticker_df["Date"])].groupby(
                "Date").nth(1).sort_values("Date", ascending=False)["Quarter Revenue"]
            # 四半期EPSのパフォーマンスを計算
            perf_previous_quarter_revenue_2 = self.calculate(
                num1=quarter_revenue, num2=quarter_revenue, num1_index=1, num2_index=5)
            perf_previous_quarter_revenue = self.calculate(
                num1=quarter_revenue, num2=quarter_revenue, num1_index=0, num2_index=4)
            perf_courrent_quarter_revenue = self.calculate(
                num1=ticker_df["CQ_Rev"].dropna(), num2=quarter_revenue, num1_index=0, num2_index=3)
            perf_next_quarter_revenue = self.calculate(
                num1=ticker_df["NQ_Rev"].dropna(), num2=quarter_revenue, num1_index=0, num2_index=2)
            self.logger.debug(
                f"`Revenue`\n2期前:{perf_previous_quarter_revenue_2}, 一期前:{perf_previous_quarter_revenue}, 今期:{perf_courrent_quarter_revenue}, 来期:{perf_next_quarter_revenue}\n")

            # 年度Revenueを設定
            annual_revenue = ticker_df[pd.notna(ticker_df["Year"])].groupby(
                "Year").nth(1).sort_values("Year", ascending=False)["Annual Revenue"]
            # 年度EPSのパフォーマンスを計算
            perf_previous_annual_revenue_2 = self.calculate(
                num1=annual_revenue, num2=annual_revenue, num1_index=1, num2_index=2)
            perf_previous_annual_revenue = self.calculate(
                num1=annual_revenue, num2=annual_revenue, num1_index=0, num2_index=1)
            perf_courrent_annual_revenue = self.calculate(
                num1=ticker_df["CY_Rev"].dropna(), num2=annual_revenue, num1_index=0, num2_index=0)
            perf_next_annual_revenue = self.calculate(
                num1=ticker_df["NY_Rev"].dropna(), num2=ticker_df["CY_Rev"].dropna(), num1_index=0, num2_index=0)
            self.logger.debug(
                f"`Revenue`\n一昨年:{perf_previous_annual_revenue_2}, 昨年:{perf_previous_annual_revenue}, 今年:{perf_courrent_annual_revenue}, 来年:{perf_next_annual_revenue}\n")

            # 各値をリスト型のdataに追加する
            data.append([ticker_code, perf_previous_quarter_eps_2, perf_previous_quarter_eps, perf_courrent_quarter_eps, perf_next_quarter_eps,
                        perf_previous_annual_eps_2, perf_previous_annual_eps, perf_courrent_annual_eps, perf_next_annual_eps,
                        perf_previous_quarter_revenue_2, perf_previous_quarter_revenue, perf_courrent_quarter_revenue, perf_next_quarter_revenue,
                        perf_previous_annual_revenue_2, perf_previous_annual_revenue, perf_courrent_annual_revenue, perf_next_annual_revenue,
                         ])

        col = ["Ticker", "Previous Quarter EPS2", "Previous Quarter EPS", "Current Quarter EPS", "Next Quarter EPS",
               "Previous Annual EPS2", "Previous Annual EPS", "Current Annual EPS", "Next Annual EPS",
               "Previous Quarter Revenue2", "Previous Quarter Revenue", "Current Quarter Revenue", "Next Quarter Revenue",
               "Previous Annual Revenue2", "Previous Annual Revenue", "Current Annual Revenue", "Next Annual Revenue"]
        append_df = pd.DataFrame(data, columns=col)

        df = pd.merge(df, append_df, on='Ticker', how='inner')

        outputDir = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "CurrentAnnual.csv")
        df.to_csv(outputDir, index=False)
        self.logger.info(f"---Done Process CurrentAnnual (CurrentAnnual)---")


"""
機関投資家の増加数を取得
"""


class Institutional:
    # 関数の定義
    def __init__(self, logger, driver):
        self.logger = logger
        self.driver = driver
    # 機関投資家数のパフォーマンスを取得する関数

    def get_inst_perf(self, i, ticker):
        if '-' in ticker:
            ticker = ticker.replace('-', '.')
        try:
            # ターゲットのURL
            load_dotenv()  # .envファイルから値を読み込む
            url = os.getenv('institute_url')
            # サイトのホームへ移動
            self.driver.get(url)
            # 各銘柄のページへ移動
            input_search = self.driver.find_element(
                By.XPATH, "/html/body/div[1]/div/div/div/div[1]/div/header/div/div[2]/div[1]/div[2]/div/div/div/div[2]/input[1]")
            input_search.clear()
            input_search.send_keys(ticker)
            input_search.send_keys(Keys.ENTER)  # エンターキーを押下
            tmp_ticker = ticker + " "  # 銘柄の検索に空白が必要なので一時的に空白を追加
            xpath_expression = f"//td[text()='{tmp_ticker}']/preceding-sibling::td/a"
            ticker_btn = self.driver.find_element(By.XPATH, xpath_expression)
            ticker_btn.click()
            # ページソースを収録
            WebDriverWait(self.driver, 15).until(EC.presence_of_element_located(
                (By.XPATH, '//*[contains(text(), "Funds Holding:")]')))
            page_source = self.driver.page_source
            soup = BeautifulSoup(page_source, "html.parser")
            # 目的の要素を取得する
            try:
                perf_inst = soup.find(string='Funds Holding:').find_parent(
                    'tr').find(class_='label').text
                # try:
                #     get_ticker = soup.find(class_="profile-default").find("h1").find("a").text
                #     true_or_false = ticker == get_ticker
                # except:
                #     true_or_false = np.nan
                #     self.logger.info("cound not be earned True or False")
            except Exception as e:
                perf_inst = np.nan
                # true_or_false = np.nan
                if self.add_error_flg == True:
                    self.error_list.append(ticker)
                self.logger.warning(
                    f"{i+1} | {ticker} Value could not be earned \n {e}")
                self.logger.warning(
                    "-------------------------------------------------------------------------------------------------------------------")
        except Exception as e:
            try:
                # 該当銘柄の企業名で再度検索する
                # ターゲットのURL
                load_dotenv()  # .envファイルから値を読み込む
                url = os.getenv('institute_url')
                self.driver.get(url)
                input_search = self.driver.find_element(
                    By.XPATH, "/html/body/div[1]/div/div/div/div[1]/div/header/div/div[2]/div[1]/div[2]/div/div/div/div[2]/input[1]")
                input_search.clear()
                company = stock_df[stock_df["Ticker"] ==
                                   ticker]["Company"].to_string(index=False)
                input_search.send_keys(company)
                input_search.send_keys(Keys.ENTER)  # エンターキーを押下
                xpath_expression = f"//td[text()='{tmp_ticker}']/preceding-sibling::td/a"
                self.driver.find_element(By.XPATH, xpath_expression).click()
                # ページソースを収録
                WebDriverWait(self.river, 15).until(EC.presence_of_element_located(
                    (By.XPATH, '//*[contains(text(), "Funds Holding:")]')))
                page_source = self.driver.page_source
                soup = BeautifulSoup(page_source, "html.parser")
                # 目的の要素を取得する
                try:
                    perf_inst = soup.find(string='Funds Holding:').find_parent(
                        'tr').find(class_='label').text
                    # try:
                    #     get_ticker = soup.find(class_="profile-default").find("h1").find("a").text
                    #     true_or_false = ticker == get_ticker
                    # except:
                    #     true_or_false = np.nan
                    #     self.logger.info("cound not be earned True or False")
                except Exception as e:
                    perf_inst = np.nan
                    # true_or_false = np.nan
                    if self.add_error_flg == True:
                        self.error_list.append(ticker)
                    self.logger.warning(
                        f"{i+1} | {ticker} Value could not be earned \n {e}")
                    self.logger.warning(
                        "-------------------------------------------------------------------------------------------------------------------")
            except:
                perf_inst = np.nan
                # true_or_false = np.nan
                if self.add_error_flg == True:
                    self.error_list.append(ticker)
                self.logger.warning(
                    f"{i+1} | {ticker} Value could not be earned (Error) \n {e}")
                self.logger.warning(
                    "-------------------------------------------------------------------------------------------------------------------")

        if self.add_error_flg == True:
            # self.data.append([ticker, perf_inst, true_or_false])
            self.data.append([ticker, perf_inst])
        else:
            # 既存の要素を上書き
            for item in self.data:
                if item[0] == ticker:
                    item[1] = perf_inst
                    # item[2] = true_or_false

    def process(self):
        # 銘柄のリストを読み込む
        stock_df = pd.read_csv(os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "BuyingStock.csv"))
        stock_df_ticker = stock_df["Ticker"]
        # リスト型のデータ変数
        self.data = []
        self.error_list = []
        # エラーが起こった際に再実行するためのリストに追加するフラグ
        self.add_error_flg = True

        # 対象リンク
        load_dotenv()  # .envファイルから値を読み込む
        base_url = os.getenv('institute_url')
        loggin_url = base_url + "/login"
        # ユーザーネーム
        username = os.getenv("institute_username")
        # パスワード
        password = os.getenv("institute_password")
        # ドライバの設定
        # txtファイルを用いて処理の実行を100銘柄づつに分ける
        remove_flg = False
        process_txt_file = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "Institutional.txt")
        is_file = os.path.isfile(process_txt_file)

        if is_file:  # ファイルが存在する場合の処理
            f = open(process_txt_file, 'r')
            start_num = int(f.read())+100
            if start_num+100 <= len(stock_df_ticker):
                end_num = start_num+100
            else:
                end_num = len(stock_df_ticker)
                remove_flg = True
        else:  # ファイルが存在しない場合の処理
            if 100 <= len(stock_df_ticker):
                start_num = 0
                end_num = 100
            else:
                start_num = 0
                end_num = len(stock_df_ticker)
                remove_flg = True
        # ファイルの書き込み
        f = open(process_txt_file, 'w')
        f.write(f"{start_num}")
        f.close()
        # ファイルの削除
        if remove_flg:
            os.remove(process_txt_file)

        # サイトにログインする
        # TODO ログイン処理時にエラーが発生する
        """
        エラーメッセージ
        urllib3.exceptions.MaxRetryError: HTTPConnectionPool(host='localhost', port=55441):
        Max retries exceeded with url: /session/4ab522c13a3fa49c29ee791db6831ca7/url
        (Caused by NewConnectionError('<urllib3.connection.HTTPConnection object at 0x11f27a310>:
        Failed to establish a new connection: [Errno 61] Connection refused'))
        """
        self.driver.get(loggin_url)  # URLを開く
        time.sleep(1.0)
        self.logger.info(self.driver.current_url)
        WebDriverWait(self.driver, 15).until(EC.presence_of_element_located(
            (By.ID, "lnk-login")))  # ページ上の指定の要素が読み込まれるまで待機（15秒でタイムアウト判定）
        self.driver.find_element(By.ID, "lnk-login").click()
        input_login = self.driver.find_element(By.ID, "login")
        input_login.clear()
        input_login.send_keys(username)
        input_pass = self.driver.find_element(By.ID, "password")
        input_pass.clear()
        input_pass.send_keys(password)
        self.driver.find_element(By.CLASS_NAME, "login").click()
        time.sleep(5)  # 待機時間
        self.logger.info(self.driver.current_url)
        load_dotenv()  # .envファイルから値を読み込む
        url = os.getenv('institute_url')
        self.driver.get(url)  # URLを開く
        time.sleep(0.5)
        # 検索をStock Onlyに変更
        self.driver.find_element(
            By.XPATH, "/html/body/div[1]/div/div/div[1]/div[1]/div/header/div/div[2]/div[1]/div[1]/div/div/div/div[1]/div[1]/div").click()
        self.driver.find_element(
            By.XPATH, "/html/body/div[1]/div/div/div[2]/div/div[4]/div/div").click()
        for i, ticker in enumerate(stock_df_ticker[start_num:end_num]):
            # 値の取得
            self.get_inst_perf(i, ticker)
            print(f"{i+start_num+1}/{end_num} ", end="\r")
            time.sleep(1.0)  # 待機時間

        # errorになった銘柄のデータを再度取得する
        if len(self.error_list) != 0:
            self.logger.info("error_list Process")
            time.sleep(15)  # 待機時間
            self.add_error_flg = False
            self.logger.info(self.error_list)
            # Errorになった銘柄のデータを再度取得する
            for i, ticker in enumerate(self.error_list):
                self.get_inst_perf(i, ticker)
                print(
                    f"{i+start_num+1}/{len(self.error_list)} ", end="\r")
                time.sleep(1.0)  # 待機時間

        try:
            # dataをdfに変換する
            df = pd.DataFrame(self.data)
            # col = ["Ticker", "Perf Inst", "True or False"]
            col = ["Ticker", "Perf Inst"]
            df.columns = col

            # dfを出力する
            outputDir = os.path.join(
                glob.glob(os.getcwd()+"/CsvData*")[0], f"PerfInst_{start_num}.csv")
            df.to_csv(outputDir, index=False)
        except Exception as e:
            self.logger.warning(e)

        # Chrome driverを終了する
        self.driver.quit()

    def execute(self):
        # 実行時間の計測
        start = time.time()
        # ファイルが存在しない場合の初回実行
        self.process()
        # ファイルが削除されるまで処理を繰り返す
        path = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "Institutional.txt")
        while os.path.exists(path):
            self.process()

        df_dict = {}
        df_dir_ = glob.glob(os.getcwd()+"/CsvData*/PerfInst_*.csv")
        df_dir = sorted(df_dir_, key=lambda x: int(
            re.search(r'\d+', x).group()))  # 数値部分を抽出してソート

        # CSVファイルを読み込んでデータフレームに結合
        dfs = [pd.read_csv(file) for file in df_dir]
        combined_df = pd.concat(dfs, ignore_index=True)

        outputDir = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "PerfInst.csv")
        combined_df.to_csv(outputDir, index=False)

        # ファイルを削除
        for file in df_dir:
            os.remove(file)

        self.logger.info(f"Time : {round(time.time() - start, 2)}")
        self.logger.info(f"---Done Process Institutional---")


"""
BuyingStock.csvに値を追加する
"""


class AppendData:
    def __init__(self, logger):
        self.logger = logger

    def execute(self):
        # BuyingStock.csv
        df_dir = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "BuyingStock.csv")
        df = pd.read_csv(df_dir)
        # CurrentAnnual.csv
        courrent_annual_df_dir = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "CurrentAnnual.csv")
        courrent_annual_df = pd.read_csv(courrent_annual_df_dir)
        # PerfInst.csv
        perf_inst_df_dir = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "PerfInst.csv")
        perf_inst_df = pd.read_csv(perf_inst_df_dir)
        # IXIC.csvとパフォーマンス
        ixic_dir = os.path.join(
            glob.glob(os.getcwd()+"/StockData*")[0], "^IXIC.csv")
        ixic_df = pd.read_csv(ixic_dir)
        index_adjclose = ixic_df['Adj Close']
        recent_index_performance = round(
            ((index_adjclose.iloc[-1] - index_adjclose.iloc[-2]) / abs(index_adjclose.iloc[-2])) * 100, 2)
        before_index_performance = round(
            ((index_adjclose.iloc[-1] - index_adjclose.iloc[-51]) / abs(index_adjclose.iloc[-51])) * 100, 2)

        for ticker_code in df["Ticker"]:
            # 各銘柄のEPSとRevenueのパフォーマンス
            condition = df["Ticker"] == ticker_code
            try:
                courrent_annual = courrent_annual_df[courrent_annual_df["Ticker"]
                                                     == ticker_code].iloc[0, :]
                # 四半期EPS
                df.loc[condition, "Previous Quarter EPS2"] = courrent_annual["Previous Quarter EPS2"]
                df.loc[condition, "Previous Quarter EPS"] = courrent_annual["Previous Quarter EPS"]
                df.loc[condition,
                       "Current Quarter EPS"] = courrent_annual["Current Quarter EPS"]
                df.loc[condition,
                       "Next Quarter EPS"] = courrent_annual["Next Quarter EPS"]
                # 年間EPS
                df.loc[condition, "Previous Annual EPS2"] = courrent_annual["Previous Annual EPS2"]
                df.loc[condition,
                       "Previous Annual EPS"] = courrent_annual["Previous Annual EPS"]
                df.loc[condition,
                       "Current Annual EPS"] = courrent_annual["Current Annual EPS"]
                df.loc[condition,
                       "Next Annual EPS"] = courrent_annual["Next Annual EPS"]
                # 四半期Revenue
                df.loc[condition, "Previous Quarter Revenue2"] = courrent_annual["Previous Quarter Revenue2"]
                df.loc[condition, "Previous Quarter Revenue"] = courrent_annual["Previous Quarter Revenue"]
                df.loc[condition, "Current Quarter Revenue"] = courrent_annual["Current Quarter Revenue"]
                df.loc[condition, "Next Quarter Revenue"] = courrent_annual["Next Quarter Revenue"]
                # 年間Revenue
                df.loc[condition, "Previous Annual Revenue2"] = courrent_annual["Previous Annual Revenue2"]
                df.loc[condition, "Previous Annual Revenue"] = courrent_annual["Previous Annual Revenue"]
                df.loc[condition, "Current Annual Revenue"] = courrent_annual["Current Annual Revenue"]
                df.loc[condition,
                       "Next Annual Revenue"] = courrent_annual["Next Annual Revenue"]
            except:
                # 四半期EPS
                df.loc[condition, "Previous Quarter EPS2"] = np.nan
                df.loc[condition, "Previous Quarter EPS"] = np.nan
                df.loc[condition, "Current Quarter EPS"] = np.nan
                df.loc[condition, "Next Quarter EPS"] = np.nan
                # 年間EPS
                df.loc[condition, "Previous Annual EPS2"] = np.nan
                df.loc[condition, "Previous Annual EPS"] = np.nan
                df.loc[condition, "Current Annual EPS"] = np.nan
                df.loc[condition, "Next Annual EPS"] = np.nan
                # 四半期Revenue
                df.loc[condition, "Previous Quarter Revenue2"] = np.nan
                df.loc[condition, "Previous Quarter Revenue"] = np.nan
                df.loc[condition, "Current Quarter Revenue"] = np.nan
                df.loc[condition, "Next Quarter Revenue"] = np.nan
                # 年間Revenue
                df.loc[condition, "Previous Annual Revenue2"] = np.nan
                df.loc[condition, "Previous Annual Revenue"] = np.nan
                df.loc[condition, "Current Annual Revenue"] = np.nan
                df.loc[condition, "Next Annual Revenue"] = np.nan
            # 各移動平均のパフォーマンス
            price = df.loc[condition, 'Price'].values[0]
            sma10 = df.loc[condition, 'SMA10'].values[0]
            ema21 = df.loc[condition, 'EMA21'].values[0]
            sma50 = df.loc[condition, 'SMA50'].values[0]
            df.loc[condition, 'SMA10 Gap'] = round(
                ((price - sma10) / abs(sma10))*100, 4)
            df.loc[condition, 'EMA21 Gap'] = round(
                ((price - ema21) / abs(ema21))*100, 2)
            df.loc[condition, 'SMA50 Gap'] = round(
                ((price - sma50) / abs(sma50))*100, 2)
            volume = float(
                df.loc[condition, 'Volume'].values[0].replace(',', ''))
            volume_avg = float(df.loc[condition, 'Volume SMA50'].values[0])
            df.loc[condition, 'Volume SMA50 Gap'] = round(
                ((volume - volume_avg) / abs(volume_avg)) * 100, 2)

            # 機関投資家増加数のパフォーマンス
            perf_inst_df = pd.read_csv(perf_inst_df_dir)
            perf_inst = perf_inst_df[perf_inst_df["Ticker"] == ticker_code]
            df.loc[df["Ticker"] == ticker_code,
                   "Perf Inst"] = perf_inst["Perf Inst"]

            # インデックスと個別銘柄を比較
            ticker_dir = glob.glob(
                os.getcwd()+f"/StockData*/{ticker_code}.csv")[0]
            ticker_df = pd.read_csv(ticker_dir)
            adjclose = ticker_df["Adj Close"]

            recent_performance = round(
                ((adjclose.iloc[-1] - adjclose.iloc[-2]) / abs(adjclose.iloc[-2])) * 100, 2)
            recent_diff_performance = round(
                (recent_performance - recent_index_performance) / abs(recent_index_performance), 2)
            before_performance = round(
                ((adjclose.iloc[-1] - adjclose.iloc[-51]) / abs(adjclose.iloc[-51])) * 100, 2)
            before_diff_performance = round(
                (before_performance - before_index_performance) / abs(before_index_performance), 2)
            df.loc[df["Ticker"] == ticker_code,
                   "Diff Index Recent"] = recent_diff_performance
            df.loc[df["Ticker"] == ticker_code,
                   "Diff Index 50days"] = before_diff_performance

        # 列の順番を入れ替え
        df = df.reindex(columns=[
            'Ticker', 'Price', 'Company', 'Sector',  'Industry', 'MarketCap', 'EarningsDate', 'Volume', 'Total', 'Prev Total', 'BuyingFlg', 'Diff Index Recent', 'Diff Index 50days',
            'Previous Quarter EPS2', 'Previous Quarter Revenue2', 'Previous Quarter EPS', 'Previous Quarter Revenue', 'Current Quarter EPS', 'Current Quarter Revenue', 'Next Quarter EPS', 'Next Quarter Revenue',  # C
            'Previous Annual EPS2', 'Previous Annual Revenue2', 'Previous Annual EPS', 'Previous Annual Revenue', 'Current Annual EPS', 'Current Annual Revenue', 'Next Annual EPS', 'Next Annual Revenue', 'ROE',  # A
            'ATH', '52W High',  # N
            'InsiderOwn', 'U/D',  # S
            'RS',  # L
            'Perf Inst',  # I
            'EMA8 Gap', 'SMA10 Gap', 'EMA21 Gap', 'SMA50 Gap', 'SMA200 Gap', 'Volume SMA50 Gap',  # 移動平均
            # 'Date', 'P/E', 'FwdP/E', 'No1', 'No4', 'No5', 'No6', '^IXIC Total', 'BuyFlg1', 'BuyFlg2', 'BuyFlg4', 'EMA8', 'SMA10', 'EMA21', 'SMA50', 'SMA200', 'Volume SMA50', # 不要な列
            'Date', 'No1', 'No4', 'No5', 'No6', '^IXIC Total', 'BuyFlg1', 'BuyFlg2', 'BuyFlg4', 'EMA8', 'SMA10', 'EMA21', 'SMA50', 'SMA200', 'Volume SMA50',  # 不要な列
        ])
        df.to_csv(df_dir, index=False)


"""
Excelに変換後視覚情報を調整
"""


class ConvertExcel:
    def __init__(self, logger):
        self.logger = logger

    # 桁数を変換
    def convert_value(self, value):
        if (isinstance(value, float)) or (pd.isna(value)):
            return value
        elif 'k' in value:
            return float(value.replace('k', '')) * 1000
        elif 'M' in value:
            return float(value.replace('M', '')) * 1000000
        elif 'B' in value:
            return float(value.replace('B', '')) * 1000000000
        else:
            return None

    # RGBの配列を16進数に変換
    def rgb_to_hex(self, rgb):
        hex_color = [round(x * 255) for x in rgb[:3]]
        hex_color = [format(x, '02X') for x in hex_color]
        hex_code = ''.join(hex_color)
        return hex_code

    # 与えられた日付の配列から1ヶ月以内の日付を判別
    def is_within_one_month(self, date_str):
        if date_str == "-":
            return False
        # 月の辞書を定義
        month_dict = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                      'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}

        # 今日の日付を取得
        now = datetime.datetime.now().strftime("%m-%d")
        # 1ヶ月後の日付を取得
        next_one_month = (datetime.datetime.now() +
                          timedelta(days=30)).strftime("%m-%d")
        # 与えられた月の値
        month = month_dict.get(date_str[:3])
        # 与えられた日の値
        date = int(date_str[4:6])
        match = re.search(r'\d{6}', date_str)
        date = int(date_str[4:6])
        # 与えられた日付
        month_date = datetime.datetime(
            datetime.datetime.now().year, month, date).strftime("%m-%d")
        return (now <= month_date) and (month_date <= next_one_month)

    # ヒートマップ作成の初期実行
    def init_heatmap(self, column_name, cmap, vmax, vmin, vcenter):
        self.df[column_name] = pd.to_numeric(
            self.df[column_name], errors="coerce")
        value = self.df[[column_name]].values
        cmap = cm.get_cmap(cmap)
        data = value.flatten()
        norm = TwoSlopeNorm(vmax=vmax, vmin=vmin, vcenter=vcenter)
        # norm = LogNorm(vmax=vmax, vmin=vmin)
        rgba_values = cmap(norm(data))
        hex_colors = [self.rgb_to_hex(rgb) for rgb in rgba_values]
        return hex_colors

    # 列の番号を列の名前から取得する
    def search_col_num(self, col_name):
        for i in range(self.ws.max_column-1):
            cell = self.ws.cell(row=1, column=i+2).value
            if cell == col_name:
                self.col_num = i+2
                break
        return self.col_num

    # Table設定
    def num2alpha(self, num):
        if num <= 26:
            return chr(64+num)
        elif num % 26 == 0:
            return self.num2alpha(num//26-1)+chr(90)
        else:
            return self.num2alpha(num//26)+chr(64+num % 26)

    def execute(self):
        # csvファイルをxlsxファイルに変換
        df_dir = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "BuyingStock.csv")
        self.df = pd.read_csv(df_dir)
        self.df.to_excel(os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "BuyingStock.xlsx"))
        df_xlsx_dir = os.path.join(
            glob.glob(os.getcwd()+"/CsvData*")[0], "BuyingStock.xlsx")

        # xlsxファイルを読み込み
        wb = openpyxl.load_workbook(df_xlsx_dir)
        self.ws = wb['Sheet1']

        # Ticker 今期EPS、今期収益、来期EPS、来期収益がすべて25%以上
        ticker_col_num = self.search_col_num(col_name='Ticker')  # Ticker列の列番号
        fill_color = PatternFill(
            fgColor='88bfbf', bgColor='88bfbf', fill_type='solid')
        # ファンダメンタルズに使用する列のみのdf
        fundamental_list = ['Current Quarter EPS', 'Current Quarter Revenue',
                            'Next Quarter EPS', 'Next Quarter Revenue']  # 対象列
        fundamental_df = self.df.loc[:, fundamental_list]
        # すべての列を判定
        for i in range(len(fundamental_df)):
            cell = self.ws.cell(row=i+2, column=ticker_col_num)
            # すべての列が25以上の場合のみ色を変える
            if (fundamental_df.iloc[i] >= 25).all():
                cell.fill = fill_color

        # InsiderOwn
        col_num = self.search_col_num(col_name='MarketCap')
        col_num_2 = self.search_col_num(col_name='InsiderOwn')
        for i in range(self.ws.max_row-1):
            cell_market_cap = self.ws.cell(row=i+2, column=col_num)
            cell_insider_own = self.ws.cell(row=i+2, column=col_num_2)
            market_cap = self.convert_value(cell_market_cap.value)
            try:
                insider_own = float(cell_insider_own.value.replace("%", ""))
            except Exception as e:
                self.logger.warning(e)
                pass
            if market_cap > 10e+9:  # 大型株を判別
                if insider_own >= 1:  # 該当している場合は背景を緑色に変更
                    fill_color = PatternFill(
                        fgColor='48B068', bgColor='48B068', fill_type='solid')
                    cell_insider_own.fill = fill_color
            else:  # その他は中小型株
                if insider_own >= 3:  # 該当している場合は背景を緑色に変更
                    fill_color = PatternFill(
                        fgColor='48B068', bgColor='48B068', fill_type='solid')
                    cell_insider_own.fill = fill_color

        # EarningsDate
        col_num = self.search_col_num(col_name='EarningsDate')
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            earnings_date = cell.value
            if self.is_within_one_month(earnings_date):
                fill_color = PatternFill(
                    fgColor='fff352', bgColor='fff352', fill_type='solid')
                cell.fill = fill_color

        # RS
        col_num = self.search_col_num(col_name='RS')
        hex_colors = self.init_heatmap(
            column_name='RS', cmap='YlGn', vmax=None, vmin=None, vcenter=0)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            fill_color = PatternFill(
                fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
            cell.fill = fill_color

        # ATH
        col_num = self.search_col_num(col_name='Price')
        col_num_2 = self.search_col_num(col_name='ATH')
        for i in range(self.ws.max_row-1):
            cell_price = self.ws.cell(row=i+2, column=col_num)
            cell_ath = self.ws.cell(row=i+2, column=col_num_2)
            price = format(float(cell_price.value), '.2f')
            ath = format(float(cell_ath.value), '.2f')
            if ath == price:
                fill_color = PatternFill(
                    fgColor='48B068', bgColor='48B068', fill_type='solid')
                cell_ath.fill = fill_color

        # 52W High
        col_num = self.search_col_num(col_name='Price')
        col_num_2 = self.search_col_num(col_name='52W High')
        for i in range(self.ws.max_row-1):
            cell_price = self.ws.cell(row=i+2, column=col_num)
            cell_w52high = self.ws.cell(row=i+2, column=col_num_2)
            price = format(float(cell_price.value), '.2f')
            w52high = format(float(cell_w52high.value), '.2f')
            if w52high == price:
                fill_color = PatternFill(
                    fgColor='48B068', bgColor='48B068', fill_type='solid')
                cell_w52high.fill = fill_color

        # EMA8
        col_num = self.search_col_num(col_name='Price')
        col_num_2 = self.search_col_num(col_name='EMA8 Gap')
        hex_colors = self.init_heatmap(
            column_name='EMA8 Gap', cmap='RdYlGn', vmax=5, vmin=-5, vcenter=0)
        for i in range(self.ws.max_row-1):
            cell_price = self.ws.cell(row=i+2, column=col_num)
            cell_price_avg = self.ws.cell(row=i+2, column=col_num_2)
            price = float(cell_price.value)
            price_avg = float(format(cell_price_avg.value, '.2f'))
            if price >= price_avg:
                fill_color = PatternFill(
                    fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
                cell_price_avg.fill = fill_color

        # SMA10
        col_num = self.search_col_num(col_name='Price')
        col_num_2 = self.search_col_num(col_name='SMA10 Gap')
        hex_colors = self.init_heatmap(
            column_name='SMA10 Gap', cmap='RdYlGn', vmax=10, vmin=-5, vcenter=0)
        for i in range(self.ws.max_row-1):
            cell_price = self.ws.cell(row=i+2, column=col_num)
            cell_price_avg = self.ws.cell(row=i+2, column=col_num_2)
            if (cell_price.value != None) & (cell_price_avg.value != None):
                price = float(cell_price.value)
                price_avg = float(cell_price_avg.value)
                if price >= price_avg:
                    fill_color = PatternFill(
                        fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
                    cell_price_avg.fill = fill_color

        # EMA21
        col_num = self.search_col_num(col_name='Price')
        col_num_2 = self.search_col_num(col_name='EMA21 Gap')
        hex_colors = self.init_heatmap(
            column_name='EMA21 Gap', cmap='RdYlGn', vmax=20, vmin=-5, vcenter=0)
        for i in range(self.ws.max_row-1):
            cell_price = self.ws.cell(row=i+2, column=col_num)
            cell_price_avg = self.ws.cell(row=i+2, column=col_num_2)
            if (cell_price.value != None) & (cell_price_avg.value != None):
                price = float(cell_price.value)
                price_avg = float(cell_price_avg.value)
                if price >= price_avg:
                    fill_color = PatternFill(
                        fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
                    cell_price_avg.fill = fill_color

        # SMA50
        col_num = self.search_col_num(col_name='Price')
        col_num_2 = self.search_col_num(col_name='SMA50 Gap')
        hex_colors = self.init_heatmap(
            column_name='SMA50 Gap', cmap='RdYlGn', vmax=50, vmin=-5, vcenter=0)
        for i in range(self.ws.max_row-1):
            cell_price = self.ws.cell(row=i+2, column=col_num)
            cell_price_avg = self.ws.cell(row=i+2, column=col_num_2)
            if (cell_price.value != None) & (cell_price_avg.value != None):
                price = float(cell_price.value)
                price_avg = float(cell_price_avg.value)
                if price >= price_avg:
                    fill_color = PatternFill(
                        fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
                    cell_price_avg.fill = fill_color

        # SMA200
        col_num = self.search_col_num(col_name='SMA200 Gap')
        hex_colors = self.init_heatmap(
            column_name='SMA200 Gap', cmap='YlGn', vmax=50, vmin=10, vcenter=30)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                sma200_gap = float(cell.value)
                # if sma200_gap >= 1.1: # 10%の閾値は検討する
                if sma200_gap >= 10:  # 10%の閾値は検討する
                    fill_color = PatternFill(
                        fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Volume SMA50
        col_num = self.search_col_num(col_name='Volume SMA50 Gap')
        hex_colors = self.init_heatmap(
            column_name='Volume SMA50 Gap', cmap='RdYlGn', vmax=50, vmin=-50, vcenter=0)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                fill_color = PatternFill(
                    fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
                cell.fill = fill_color

        # Previous Quarter EPS2
        col_num = self.search_col_num(col_name='Previous Quarter EPS2')
        green_hex_colors = self.init_heatmap(
            column_name='Previous Quarter EPS2', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Previous Quarter EPS2', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                if float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Previous Quarter EPS
        col_num = self.search_col_num(col_name='Previous Quarter EPS')
        green_hex_colors = self.init_heatmap(
            column_name='Previous Quarter EPS', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Previous Quarter EPS', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Current Quarter EPS
        col_num = self.search_col_num(col_name='Current Quarter EPS')
        green_hex_colors = self.init_heatmap(
            column_name='Current Quarter EPS', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Current Quarter EPS', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Next Quarter EPS
        col_num = self.search_col_num(col_name='Next Quarter EPS')
        green_hex_colors = self.init_heatmap(
            column_name='Next Quarter EPS', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Next Quarter EPS', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Previous Annual EPS2
        col_num = self.search_col_num(col_name='Previous Annual EPS2')
        green_hex_colors = self.init_heatmap(
            column_name='Previous Annual EPS2', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Previous Annual EPS2', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Previous Annual EPS
        col_num = self.search_col_num(col_name='Previous Annual EPS')
        green_hex_colors = self.init_heatmap(
            column_name='Previous Annual EPS', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Previous Annual EPS', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Current Annual EPS
        col_num = self.search_col_num(col_name='Current Annual EPS')
        green_hex_colors = self.init_heatmap(
            column_name='Current Annual EPS', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Current Annual EPS', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Next Annual EPS
        col_num = self.search_col_num(col_name='Next Annual EPS')
        green_hex_colors = self.init_heatmap(
            column_name='Next Annual EPS', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Next Annual EPS', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Previous Quarter Revenue2
        col_num = self.search_col_num(col_name='Previous Quarter Revenue2')
        green_hex_colors = self.init_heatmap(
            column_name='Previous Quarter Revenue2', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Previous Quarter Revenue2', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Previous Quarter Revenue
        col_num = self.search_col_num(col_name='Previous Quarter Revenue')
        green_hex_colors = self.init_heatmap(
            column_name='Previous Quarter Revenue', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Previous Quarter Revenue', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Current Quarter Revenue
        col_num = self.search_col_num(col_name='Current Quarter Revenue')
        green_hex_colors = self.init_heatmap(
            column_name='Current Quarter Revenue', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Current Quarter Revenue', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Next Quarter Revenue
        col_num = self.search_col_num(col_name='Next Quarter Revenue')
        green_hex_colors = self.init_heatmap(
            column_name='Next Quarter Revenue', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Next Quarter Revenue', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Previous Annual Revenue2
        col_num = self.search_col_num(col_name='Previous Annual Revenue2')
        green_hex_colors = self.init_heatmap(
            column_name='Previous Annual Revenue2', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Previous Annual Revenue2', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Previous Annual Revenue
        col_num = self.search_col_num(col_name='Previous Annual Revenue')
        green_hex_colors = self.init_heatmap(
            column_name='Previous Annual Revenue', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Previous Annual Revenue', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Current Annual Revenue
        col_num = self.search_col_num(col_name='Current Annual Revenue')
        green_hex_colors = self.init_heatmap(
            column_name='Current Annual Revenue', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Current Annual Revenue', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Next Annual Revenue
        col_num = self.search_col_num(col_name='Next Annual Revenue')
        green_hex_colors = self.init_heatmap(
            column_name='Next Annual Revenue', cmap='YlGn', vmax=100, vmin=25, vcenter=62.5)
        red_hex_colors = self.init_heatmap(
            column_name='Next Annual Revenue', cmap='OrRd_r', vmax=0, vmin=-100, vcenter=-50)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                if float(cell.value) >= 25:
                    fill_color = PatternFill(
                        fgColor=green_hex_colors[i], bgColor=green_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color
                elif float(cell.value) <= 0:
                    fill_color = PatternFill(
                        fgColor=red_hex_colors[i], bgColor=red_hex_colors[i], fill_type='solid')
                    cell.fill = fill_color

        # Perfome Institute
        col_num = self.search_col_num(col_name='Perf Inst')
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            if cell.value != None:
                perf_inst = cell.value.replace('%', '')
                if float(perf_inst) >= 0:
                    fill_color = PatternFill(
                        fgColor='48B068', bgColor='48B068', fill_type='solid')
                    cell.fill = fill_color
                else:
                    fill_color = PatternFill(
                        fgColor='E75039', bgColor='E75039', fill_type='solid')
                    cell.fill = fill_color

        # Diff Index Recent
        col_num = self.search_col_num(col_name='Diff Index Recent')
        hex_colors = self.init_heatmap(
            column_name='Diff Index Recent', cmap='RdYlGn', vmax=5, vmin=-5, vcenter=0)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            fill_color = PatternFill(
                fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
            cell.fill = fill_color

        # Diff Index 50days
        col_num = self.search_col_num(col_name='Diff Index 50days')
        hex_colors = self.init_heatmap(
            column_name='Diff Index 50days', cmap='RdYlGn', vmax=5, vmin=-5, vcenter=0)
        for i in range(self.ws.max_row-1):
            cell = self.ws.cell(row=i+2, column=col_num)
            fill_color = PatternFill(
                fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
            cell.fill = fill_color

        # 不要な列を削除する
        delete_col = []
        columns = ['Date', 'No1', 'No4', 'No5', 'No6', '^IXIC Total', 'BuyFlg1', 'BuyFlg2',
                   'BuyFlg4', 'EMA8', 'SMA10', 'EMA21', 'SMA50', 'SMA200', 'Volume SMA50']
        for col in columns:
            self.col_num = self.search_col_num(col_name=col)
            delete_col.append(self.col_num)
        for col_index in sorted(delete_col, reverse=True):
            self.ws.delete_cols(col_index)

        # 一行目にインデックスを追加
        self.ws.insert_rows(1)
        self.ws.cell(row=1, column=15).value = "C"
        self.ws.cell(row=1, column=23).value = "A"
        self.ws.cell(row=1, column=31).value = "N"
        self.ws.cell(row=1, column=33).value = "S"
        self.ws.cell(row=1, column=35).value = "L"
        self.ws.cell(row=1, column=37).value = "I"
        self.ws.cell(row=1, column=38).value = "Moving Averages"
        # 文字を太字にする
        font = Font(size=14, bold=True)
        bold_list = [15, 23, 31, 33, 35, 36, 37, 38]
        for i in range(len(bold_list)):
            cell = self.ws.cell(row=1, column=bold_list[i])
            self.ws[cell.coordinate].font = font
        # セルの高さの調整
        self.ws.row_dimensions[1].height = 18
        # 罫線を引く
        side = Side(style='thick', color='000000')
        border = Border(left=side)
        border_list = [15, 19, 23, 27, 31, 33, 35, 37, 38, 43]
        max_row = self.ws.max_row+1
        for i in range(len(border_list)):
            for j in range(2, max_row):
                cell = self.ws.cell(row=j, column=border_list[i])
                self.ws[cell.coordinate].border = border
        # セルを結合
        self.ws.merge_cells('O1:V1')
        self.ws.merge_cells('W1:AD1')
        self.ws.merge_cells('AE1:AF1')
        self.ws.merge_cells('AG1:AH1')
        self.ws.merge_cells('AI1:AJ1')
        self.ws.merge_cells('AL1:AQ1')
        # セルの名称を変更
        self.ws.cell(row=2, column=15).value = "EPS 2期前"
        self.ws.cell(row=2, column=16).value = "収益 2期前"
        self.ws.cell(row=2, column=17).value = "EPS 1期前"
        self.ws.cell(row=2, column=18).value = "収益 1期前"
        self.ws.cell(row=2, column=19).value = "EPS 今期"
        self.ws.cell(row=2, column=20).value = "収益 今期"
        self.ws.cell(row=2, column=21).value = "EPS 来期"
        self.ws.cell(row=2, column=22).value = "収益 来期"
        self.ws.cell(row=2, column=23).value = "EPS 2年前"
        self.ws.cell(row=2, column=24).value = "収益 2年前"
        self.ws.cell(row=2, column=25).value = "EPS 1年前"
        self.ws.cell(row=2, column=26).value = "収益 1年前"
        self.ws.cell(row=2, column=27).value = "EPS 今年"
        self.ws.cell(row=2, column=28).value = "収益 今年"
        self.ws.cell(row=2, column=29).value = "EPS 来年"
        self.ws.cell(row=2, column=30).value = "収益 来年"
        self.ws.cell(row=2, column=38).value = "EMA8"
        self.ws.cell(row=2, column=39).value = "SMA10"
        self.ws.cell(row=2, column=40).value = "EMA21"
        self.ws.cell(row=2, column=41).value = "SMA50"
        self.ws.cell(row=2, column=42).value = "SMA200"
        self.ws.cell(row=2, column=43).value = "Volume SMA50"

        # ウィンドウ枠の固定
        self.ws.freeze_panes = 'D3'
        # オートフィルタ範囲の設定
        self.ws.auto_filter.ref = 'A2:AP2'
        # # エクセルのセル幅を自動調整
        # max_row = self.ws.max_row
        # alpha = [self.num2alpha(n) for n in range(1, self.ws.max_column+1)]
        # for c in range(len(alpha)):
        #     column = alpha[c]
        #     l = []
        #     m = 0
        #     for i in range(0, max_row, 1):
        #         l.append(len(re.sub('\.','',str(self.ws[column][i].value))))
        #         if column=='A':
        #             self.ws.column_dimensions[column].width = 14
        #         else:
        #             if column=='B' or column=='C' or column=='D':
        #                 pass
        #             else:
        #                 z = [z for z in range(1, len(l)) if l[z]>=math.floor(mean(l))*2]
        #                 y = [ l[z[y]] for y in range(len(z))]
        #                 [l.remove(y[x]) for x in range(len(y))]
        #         m = max(l)
        #         self.ws.column_dimensions[column].width = (m*1.2)+7

        # xlsxファイルの保存
        dt_now = datetime.datetime.now()
        date = dt_now.strftime('%y%m%d')
        output_data = os.path.join(
            glob.glob(os.getcwd()+"/ExcelData")[0], f"BuyingStock_{date}.xlsx")
        wb.save(output_data)
        self.logger.info(f"---Done Process ConvertExcel ---")
